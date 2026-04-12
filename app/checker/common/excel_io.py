from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def load_workbook_safe(path: str | Path) -> Workbook:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    return load_workbook(filename=p, data_only=True, read_only=False)


def sheet_as_matrix(ws: Worksheet, max_row: int | None = None, max_col: int | None = None) -> list[list[Any]]:
    rows: list[list[Any]] = []
    mr = max_row or ws.max_row
    mc = max_col or ws.max_column
    for r in range(1, mr + 1):
        row: list[Any] = []
        for c in range(1, mc + 1):
            row.append(ws.cell(row=r, column=c).value)
        rows.append(row)
    return rows


def get_cell_str(val: Any) -> str | None:
    if val is None:
        return None
    if isinstance(val, str):
        return val
    return str(val)
