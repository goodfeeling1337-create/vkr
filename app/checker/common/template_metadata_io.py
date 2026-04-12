from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from typing import Literal

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.domain.metadata import TemplateMetadata

log = logging.getLogger(__name__)

META_SHEET_NAME = "__template_meta__"
# Hidden-ish column used for redundant storage (ZZ = 702)
META_COL_START = 702
META_ROW_START = 1
META_PARTS = 5

MetadataSource = Literal["hidden_sheet", "hidden_cells", "none"]


@dataclass(frozen=True)
class MetadataReadResult:
    """Результат чтения метаданных шаблона и источник (для аудита и fallback)."""

    metadata: TemplateMetadata | None
    source: MetadataSource


def write_metadata_hidden_cells(ws: Worksheet, meta: TemplateMetadata) -> None:
    payload = json.dumps(meta.to_json_dict(), ensure_ascii=False)
    chunks: list[str] = []
    size = max(1, len(payload) // META_PARTS + 1)
    for i in range(0, len(payload), size):
        chunks.append(payload[i : i + size])
    while len(chunks) < META_PARTS:
        chunks.append("")
    for i, ch in enumerate(chunks[:META_PARTS]):
        ws.cell(row=META_ROW_START + i, column=META_COL_START, value=ch)
    ws.column_dimensions[ws.cell(row=1, column=META_COL_START).column_letter].hidden = True


def write_metadata_sheet(wb: Workbook, meta: TemplateMetadata) -> None:
    if META_SHEET_NAME in wb.sheetnames:
        del wb[META_SHEET_NAME]
    ws = wb.create_sheet(META_SHEET_NAME)
    ws.sheet_state = "hidden"
    ws["A1"] = json.dumps(meta.to_json_dict(), ensure_ascii=False)


def _read_joined_hidden(ws: Worksheet) -> str | None:
    parts: list[str] = []
    for i in range(META_PARTS):
        v = ws.cell(row=META_ROW_START + i, column=META_COL_START).value
        if v is None:
            parts.append("")
        else:
            parts.append(str(v))
    s = "".join(parts).strip()
    return s if s else None


def read_metadata_from_workbook(wb: Workbook) -> MetadataReadResult:
    # Prefer hidden sheet (устойчивее к сдвигам столбцов)
    if META_SHEET_NAME in wb.sheetnames:
        ws = wb[META_SHEET_NAME]
        raw = ws["A1"].value
        if raw:
            try:
                meta = TemplateMetadata.from_json_str(str(raw))
                log.debug("template metadata from hidden sheet version=%s", meta.reference_version_id)
                return MetadataReadResult(meta, "hidden_sheet")
            except Exception as e:  # noqa: BLE001
                log.warning("Hidden sheet meta parse failed: %s", e)
    ws_main = wb[wb.sheetnames[0]]
    joined = _read_joined_hidden(ws_main)
    if joined:
        try:
            meta = TemplateMetadata.from_json_str(joined)
            log.debug("template metadata from hidden cells version=%s", meta.reference_version_id)
            return MetadataReadResult(meta, "hidden_cells")
        except Exception as e:  # noqa: BLE001
            log.warning("Hidden cells meta parse failed: %s", e)
    return MetadataReadResult(None, "none")


def merge_metadata_preference(a: TemplateMetadata | None, b: TemplateMetadata | None) -> TemplateMetadata | None:
    """If both exist and mismatch, prefer higher template_version."""
    if a is None:
        return b
    if b is None:
        return a
    if a.reference_version_id != b.reference_version_id:
        return b if b.template_version >= a.template_version else a
    return b if b.template_version >= a.template_version else a
