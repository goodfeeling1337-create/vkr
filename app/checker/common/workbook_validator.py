from __future__ import annotations

from dataclasses import dataclass, field

from openpyxl.workbook.workbook import Workbook

from app.checker.common.section_locator import SectionLocator

META_SHEET_NAME = "__template_meta__"


def data_sheet_names(wb: Workbook) -> list[str]:
    """Имена листов с ответами (без служебного листа метаданных шаблона)."""
    return [n for n in wb.sheetnames if n != META_SHEET_NAME]


@dataclass
class WorkbookValidationResult:
    ok: bool
    errors: list[str]
    sheet_name: str | None = None
    warnings: list[str] = field(default_factory=list)


class WorkbookValidator:
    """Validates structural constraints for .xlsx homework files."""

    def validate(self, wb: Workbook, *, require_all_sections: bool = False) -> WorkbookValidationResult:
        errors: list[str] = []
        warnings: list[str] = []
        if wb is None:
            return WorkbookValidationResult(False, ["Файл не загружен"], None, [])
        names = data_sheet_names(wb)
        if len(names) != 1:
            extra = len(wb.sheetnames) - len(names)
            if extra and len(names) == 0:
                errors.append("В книге только служебные листы; нужен лист с заданиями")
            else:
                errors.append(
                    f"Ожидается один лист с ответами (плюс опционально «{META_SHEET_NAME}»), "
                    f"листов с заданиями: {len(names)}",
                )
            return WorkbookValidationResult(False, errors, None, [])
        sheet_name = names[0]
        ws = wb[sheet_name]
        matrix: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            matrix.append(list(row))
        loc = SectionLocator(matrix)
        hits = loc.find_sections()
        for n in range(1, 14):
            if n not in hits:
                msg = (
                    f"Не найдена секция задания {n} (ожидается ячейка вида «Задание №{n}» или «Задание {n}»)"
                )
                if require_all_sections:
                    errors.append(msg)
                else:
                    warnings.append(msg)
        ok = len(errors) == 0
        return WorkbookValidationResult(ok, errors, sheet_name, warnings)
