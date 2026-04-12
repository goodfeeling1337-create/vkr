from __future__ import annotations

import logging
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from app.checker.common.section_locator import SectionLocator
from app.checker.common.template_metadata_io import (
    META_SHEET_NAME,
    write_metadata_hidden_cells,
    write_metadata_sheet,
)
from app.domain.metadata import TemplateMetadata

log = logging.getLogger(__name__)


def _clear_region(ws: Any, r1: int, r2: int, c1: int, c2: int) -> None:
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.value = None


def build_student_template(
    source_path: Path,
    dest_path: Path,
    meta: TemplateMetadata,
) -> None:
    """Copy reference workbook, clear answer bodies, embed metadata."""
    shutil.copy2(source_path, dest_path)
    wb = load_workbook(dest_path, read_only=False, data_only=False)
    ws = wb[wb.sheetnames[0]]
    matrix: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append(list(row))
    loc = SectionLocator(matrix)
    hits = loc.find_sections()
    ordered = sorted(hits.items(), key=lambda x: (x[1].row, x[1].col))
    max_row = ws.max_row
    max_col = ws.max_column
    for i, (_tn, hit) in enumerate(ordered):
        start = hit.row
        end_row = ordered[i + 1][1].row - 1 if i + 1 < len(ordered) else max_row
        body_r1 = start + 1
        if body_r1 <= end_row:
            _clear_region(ws, body_r1, end_row, 1, max_col)
    write_metadata_hidden_cells(ws, meta)
    write_metadata_sheet(wb, meta)
    wb.save(dest_path)
    log.info("Template written: %s", dest_path)
