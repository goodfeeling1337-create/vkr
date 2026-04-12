from io import BytesIO

import openpyxl
from openpyxl import load_workbook

from app.checker.common.workbook_validator import WorkbookValidator


def test_rejects_multi_sheet() -> None:
    wb = openpyxl.Workbook()
    wb.create_sheet("Second")
    v = WorkbookValidator().validate(wb)
    assert not v.ok


def test_accepts_answer_sheet_plus_template_meta_sheet() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    row = 1
    for n in range(1, 14):
        ws.cell(row=row, column=2, value=f"Задание №{n}")
        row += 1
    wb.create_sheet("__template_meta__")
    v = WorkbookValidator().validate(wb)
    assert v.ok
    assert v.sheet_name == "Лист1"


def test_accepts_single_sheet_with_sections() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    row = 1
    for n in range(1, 14):
        ws.cell(row=row, column=2, value=f"Задание №{n}")
        row += 1
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf, data_only=True)
    v = WorkbookValidator().validate(wb2)
    assert v.ok
    assert v.sheet_name == "Лист1"


def test_lenient_missing_sections_are_warnings_not_errors() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.cell(row=1, column=1, value="Задание №1")
    v = WorkbookValidator().validate(wb, require_all_sections=False)
    assert v.ok
    assert v.warnings
    assert any("2" in w or "задания 2" in w.lower() for w in v.warnings)

    v_strict = WorkbookValidator().validate(wb, require_all_sections=True)
    assert not v_strict.ok
    assert v_strict.errors
