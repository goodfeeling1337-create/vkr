from __future__ import annotations

from openpyxl import Workbook

from app.checker.engine import run_check


def test_run_check_applies_cross_task_checks() -> None:
    wb = Workbook()
    ws = wb.active

    # Task 1: two rows in source relation.
    ws["A1"] = "Задание №1"
    ws["A2"] = "R"
    ws["A3"] = "A"
    ws["B3"] = "B"
    ws["A4"] = "1"
    ws["B4"] = "x"
    ws["A5"] = "1"
    ws["B5"] = "y"

    # Task 3: one row in 1NF (n3 <= n1) => EXAMPLE_ALREADY_NORMALIZED cross-check.
    ws["A7"] = "Задание №3"
    ws["A8"] = "R"
    ws["A9"] = "A*"
    ws["B9"] = "B"
    ws["A10"] = "1"
    ws["B10"] = "x"

    ref_payloads = {
        1: {"relation": "R", "headers": ["A", "B"], "rows": [("1", "x"), ("1", "y")]},
        3: {"relation": "R", "headers": ["A", "B"], "key_attributes": ["A"], "rows": [("1", "x")]},
    }
    report = run_check(wb, ref_payloads, allow_optional_pure_junction=True)

    task1 = report.task_results[1]
    codes = {m.get("code") for m in (task1.typical_mistakes or [])}
    assert "EXAMPLE_ALREADY_NORMALIZED" in codes
