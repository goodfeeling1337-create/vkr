from __future__ import annotations

import io
from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture
def sample_workbook_bytes() -> bytes:
    """Minimal valid structure: one sheet, sections 1..13, simple task 1 table."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    row = 1
    for n in range(1, 14):
        ws.cell(row=row, column=2, value=f"Задание №{n}")
        row += 1
        if n == 1:
            ws.cell(row=row, column=1, value="R")
            row += 1
            ws.cell(row=row, column=1, value="A")
            ws.cell(row=row, column=2, value="B")
            row += 1
            ws.cell(row=row, column=1, value="1")
            ws.cell(row=row, column=2, value="2")
            row += 1
        elif n in (10, 12):
            ws.cell(row=row, column=1, value="Произвольный текст")
            row += 1
        elif n == 4:
            ws.cell(row=row, column=1, value="A -> B")
            row += 1
        elif n == 2:
            ws.cell(row=row, column=1, value="Группа 1:")
            row += 1
            ws.cell(row=row, column=1, value="A, B")
            row += 1
        elif n == 3:
            ws.cell(row=row, column=1, value="R1")
            row += 1
            ws.cell(row=row, column=1, value="A*")
            ws.cell(row=row, column=2, value="B")
            row += 1
            ws.cell(row=row, column=1, value="x")
            ws.cell(row=row, column=2, value="y")
            row += 1
        elif n in (5, 6, 7, 8, 9):
            ws.cell(row=row, column=1, value="A -> B")
            row += 1
        elif n in (11, 13):
            ws.cell(row=row, column=1, value="Отношение R1:")
            row += 1
            ws.cell(row=row, column=1, value="A, B")
            row += 1
            ws.cell(row=row, column=1, value="Ключ: A")
            row += 1
        else:
            row += 1
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
